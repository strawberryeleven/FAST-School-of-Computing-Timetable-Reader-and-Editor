import sys
import difflib
import pandas as pd
import pandas.core.series
import openpyxl as xl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet as xl_Worksheet

# Function to get the day from a string
def _get_day(curr_day: str, cache: dict = None) -> str:
    """Return the first chronological day present in the given
      string as a lowercase string. Return '' if none is found"""
    if curr_day is None:
        return ''

    if cache is None:
        cache = {}

    if curr_day in cache:
        return cache[curr_day]

    days = ['monday', 'tuesday', 'wednesday', 'thursday',
            'friday', 'saturday', 'sunday']

    curr_day = curr_day.lower()

    for day in days:
        if day in curr_day:
            cache[curr_day] = day.capitalize()
            return cache[curr_day]

    return ''


# Function to parse timetable from Excel sheet (main timetable sheet)
def parse_timetable(sheet: xl_Worksheet) -> pd.DataFrame:
    total_columns = sheet.max_column

    # Starting coordinates of the actual subjects' schedule
    STARTING_ROW, STARTING_COL = 5, 2

    cell_coordinates = {}
    for cell in sheet.merged_cells.ranges:
        cell_coordinates[(cell.min_row, cell.min_col)] = cell.size['columns']

    courses = []
    total_courses = 0
    course_cache = {}

    day = ''

    # Extract starting time from the sheet
    starting_time = sheet[STARTING_ROW - 1][2].value
    if starting_time is None or ':' not in starting_time:
        hours_offset = 8
        minutes_offset = 30
    else:
        starting_time = starting_time.split(':')
        hours_offset = int(starting_time[0])
        minutes_offset = int(starting_time[1][:2])
        is_pm = 'pm' in starting_time[1][2:].lower()
        if is_pm:
            hours_offset += 12

    for row in sheet.iter_rows(min_row=STARTING_ROW):
        if row[0].value is not None:
            day = _get_day(row[0].value)

        room = row[1].value
        if room is None:
            continue
        room = room.strip()

        col_no = STARTING_COL
        col_no_offset = 0
        while col_no < total_columns:
            col_letter = get_column_letter(col_no + 1)
            if sheet.column_dimensions[col_letter].hidden:
                col_no += 1
                col_no_offset += 1
                continue

            cell = row[col_no]
            if cell.value is None:
                col_no += cell_coordinates.get((cell.row, cell.column), 1)
                continue

            if '(' not in cell.value:
                col_no += cell_coordinates.get((cell.row, cell.column), 1)
                continue

            course_details = cell.value.split('(')

            # Ignore anything in brackets within the course name
            # Such things like are of no use to us
            title = course_details[0].split('(')[0].strip()
            # Replace '&' with 'and' for easier matching later on
            title = title.replace('&', 'and', 1)

            section_list = course_details[1].strip().rstrip(')').split(',')

            start_time = [(hours_offset + (col_no - col_no_offset - STARTING_COL) // 6),
                          minutes_offset + ((col_no - col_no_offset - STARTING_COL) % 6) * 10]
            if start_time[1] >= 60:
                start_time[0] += 1
                start_time[1] -= 60

            cell_length = cell_coordinates.get((cell.row, cell.column), 1)

            if cell_length == 1:
                cells_remaining = total_columns - col_no
                starting_color = cell.fill.start_color.index
                while cell_length < cells_remaining:
                    end_cell = row[col_no + cell_length]
                    if (end_cell.value is not None or
                            end_cell.fill.start_color.index != starting_color):
                        break
                    cell_length += 1
                    if cell.border.right.style is not None:
                        break

            col_no += cell_length

            length_of_class = [(cell_length // 6),
                               (cell_length % 6) * 10]
            end_time = [start_time[i] + length_of_class[i] for i in range(2)]
            if end_time[1] >= 60:
                end_time[0] += 1
                end_time[1] -= 60

            current_lectures = []

            for section in section_list:
                section = section.strip()

                # Append a new entry for each occurrence of the same course and instructor
                total_courses += 1

                courses.append({
                    'title': title,
                    'section': section,
                    'room': room,
                    'day': day,
                    'start_time': f'{start_time[0]:02}:{start_time[1]:02}',
                    'end_time': f'{end_time[0]:02}:{end_time[1]:02}',
                })

    if len(courses) > 0:
        return pd.DataFrame(courses, columns=['title', 'section', 'room', 'day', 'start_time', 'end_time'])
    return pd.DataFrame()


# Function to get course details from multiple sheets in excel (CS, SE, DS etc)
def get_course_details(workbook: xl.Workbook, sheets_list: list[xl_Worksheet]) -> pd.DataFrame:
    """Extract the course details from all sheets in 'sheets_list' within the
       open workbook. Return a pandas dataframe containing the extracted
       data."""
    course_details = []

    for sheet_name in sheets_list:
        sheet = workbook[sheet_name]

        starting_row = -1
        col_num = {}
        course_cache = set()

        for row_no in range(2, sheet.max_row):
            columns_in_sheet = []

            for cell in sheet[row_no]:
                if cell.value is None:
                    break
                if type(cell.value) is not str:
                    continue
                columns_in_sheet.append(cell.value.lower().strip())

            col_num.clear()

            # Do we even have the columns ?
            for index, col_name in enumerate(columns_in_sheet):
                col_name = col_name.lower()
                if 'code' not in col_num and 'code' in col_name:
                    col_num['code'] = index
                elif 'title' not in col_num and ('title' in col_name or
                                                 'course' in col_name):
                    col_num['title'] = index
                elif 'section' not in col_num and 'section' in col_name:
                    col_num['section'] = index
                elif ('instructor' not in col_num and
                        ('teacher' in col_name or 'instructor' in col_name)):
                    col_num['instructor'] = index
                elif ('credit_hours' not in col_num and
                        'credit hour' in col_name):
                    col_num['credit_hours'] = index
                elif 'offered_for' not in col_num and 'offered' in col_name:
                    col_num['offered_for'] = index
                elif 'category' not in col_num and 'category' in col_name:
                    col_num['category'] = index

            # Do we have our main identifiers ?
            if ('code' in col_num and 'section' in col_num and
                    'title' in col_num):
                # First row to start parsing at
                starting_row = row_no + 1
                break

        if starting_row == sheet.max_row or starting_row == -1:
            break

        repeated = False
        for row in sheet.iter_rows(min_row=starting_row, values_only=True):
            code, title = row[col_num['code']], row[col_num['title']]
            section = row[col_num['section']]
            if title is None:
                continue
            if code is None or section is None:
                if 'repeat' in title.lower():
                    repeated = True
                else:
                    repeated = False
                continue


            title = title.split('(')[0].strip().replace('&', 'and', 1)
            code = code.strip()
            section = section.strip()

            # Skip duplicates if any
            if (title, section) in course_cache:
                continue
            course_cache.add((title, section))

            course = {
                'title': title,
                'code': code,
                'section': section,
            }

            if 'instructor' in col_num:
                instructor = row[col_num['instructor']]
                if instructor is not None:
                    # Ignore VF/CC if mentioned
                    instructor = instructor.split('(')[0].strip()
                course['instructor'] = instructor

            if 'credit_hours' in col_num:
                credit_hours = row[col_num['credit_hours']]
                if type(credit_hours) is not int:
                    credit_hours = None
                course['credit_hours'] = credit_hours

            if 'offered_for' in col_num:
                offered_for = row[col_num['offered_for']]
                if offered_for is not None and type(offered_for) is str:
                    if '(' in offered_for:
                        offered_for = offered_for.split('(')
                        program = offered_for[0].strip()
                        target_dept = offered_for[1].strip().rstrip(')')
                    else:
                        program = offered_for[:2]
                        target_dept = offered_for[2:].strip()
                else:
                    program, target_dept = None, None

                course['program'] = program
                course['target_department'] = target_dept

            if 'category' in col_num:
                category = row[col_num['category']]
                if category is not None and '(' in category:
                    category = category.split('(')
                    parent_dept = category[0].strip()
                    course_type = category[1].strip().lstrip('(').rstrip(')')
                else:
                    parent_dept = _get_dept_from_course_code(code)
                    if parent_dept == '':
                        parent_dept = target_dept
                    course_type = category

                course['parent_department'] = parent_dept
                course['type'] = course_type

            course['repeat'] = repeated

            course_details.append(course)

    if len(course_details) > 0:
        return pd.DataFrame(course_details, columns=course_details[0].keys())
    return pd.DataFrame()


# Function to merge course details with timetable
def merge_timetable_with_details(course_details: pd.DataFrame, timetable: pd.DataFrame) -> pd.DataFrame:
    # Match titles between the two dataframes
    timetable['title'] = timetable.apply(
       lambda row: _get_corresponding_title(row, course_details),
       axis=1
    )

    course_data = course_details.merge(
       timetable, on=['section', 'title'], how='right',
       indicator=True
    )
    course_data['instructor'] = course_data['instructor'].fillna('');

    # Seperate the right_only data for further processing
    course_data_unmerged = course_data[
        course_data['_merge'] == 'right_only'
    ][timetable.columns]
    # Convert course_data to a inner merged one
    course_data = course_data[course_data['_merge'] == 'both']

    # No use for merge column anymore
    course_data.drop('_merge', axis=1, inplace=True)

    # Is there even a need of further processing ?
    if course_data_unmerged.empty:
        return course_data

    # Temporary column for secondary merge
    course_details['sec_tmp'] = course_details['section'] \
                                    .apply(lambda sec : sec[:6])
    course_data_unmerged['sec_tmp'] = course_data_unmerged['section'] \
                                        .apply(lambda sec: sec[:6])

    # Drop temporary & overlapping columns
    course_data_unmerged = course_data_unmerged.merge(
        course_details, on=['sec_tmp', 'title'], how='left',
        suffixes=('', '_details')
    ).drop(['section_details', 'sec_tmp'], axis=1)
    course_details = course_details.drop('sec_tmp', axis=1)
    # Append to the end of old data to convert it back to original state
    course_data = pd.concat(
        [course_data, course_data_unmerged],
        ignore_index=True
    )

    return course_data


# Function to get the parent department from the course code
def _get_dept_from_course_code(course_code: str) -> str:
    """Return the parent department corresponding to the course code.
      Return an empty string ('') for unknown course codes."""
    departments = {'NS': 'NS', 'MT': 'HSS', 'MT': 'HSS', 'MT':'HSS',
                   'CS' : 'CS', 'SE': 'CS', 'DS': 'DS'}

    if course_code is None:
        return ''
    return departments.get(course_code[:2], '')


# Function to get corresponding title from course details
def _get_corresponding_title(row: pandas.core.series.Series,
                             details_df: pd.DataFrame) -> str:
    """Return the closest matching title for the current row from the
      details DataFrame."""
    title = row['title']
    section = row['section']

    details_exact_match = details_df[
        details_df['section'].apply(lambda sec: sec == row['section'])
    ]
    res = difflib.get_close_matches(title, details_exact_match['title'])

    # Any matches already ?
    if len(res) > 0:
        return res[0]

    # Do we have a different section for the same semester?
    details_semi_match = details_df[
        details_df['section'].apply(
            lambda sec: section[:5] in sec if type(sec) is str else False)
    ]

    res = difflib.get_close_matches(title, details_semi_match['title'])

    # Any luck this time ?
    if len(res) > 0:
        return res[0]

    res = difflib.get_close_matches(title, details_df['title'])

    return res[0] if len(res) > 0 else title


# Function to write DataFrame to CSV
def write_to_csv(course_data: pd.DataFrame, output_file: str) -> None:
    course_data.to_csv(output_file, index=False)
    print(f'Data written to CSV file: {output_file}')


# Function to write DataFrame to Excel
def write_to_excel(course_data: pd.DataFrame, output_file: str) -> None:
    course_data.to_excel(output_file, index=False)
    print(f'Data written to Excel file: {output_file}')



def main():

    filename = 'FSC_Time_Table__List_of_Courses_Fall_2023_v1.1.xlsx'
    output_excel_filename = 'processedTimeTableEXCEL.xlsx'
    output_csv_filename = 'processedTimeTableCSV.csv'

    # Load the Excel file
    print(f'Attempting to open {filename}')
    try:
        workbook = xl.load_workbook(filename)
    except FileNotFoundError:
        sys.stderr.write(f'Error: Unable to open file: {filename}.\n')
        sys.exit(2)

    print(f'Successfully opened {filename}')

    list_of_sheets = workbook.sheetnames
    timetable_sheet = workbook.active.title

    for sheet in list_of_sheets:
        sheet_lower = sheet.lower()
        if 'tt' in sheet_lower or 'timetable' in sheet_lower:
            if timetable_sheet != sheet:
                timetable_sheet = sheet
            break

    list_of_sheets.remove(timetable_sheet)

    print('\nExtracting course details...')
    course_details = get_course_details(workbook, list_of_sheets)
    print('Done.')

    print('\nExtracting class details...')
    course_timetable = parse_timetable(workbook[timetable_sheet])
    print('Done.')

    # Update timetable's course titles to match those in course details
    print('\nMerging course and class details...')
    course_data = merge_timetable_with_details(
        course_details=course_details,
        timetable=course_timetable,
    )
    print('Done.')

    # Write to Excel
    print(f'\nGenerating {output_excel_filename}')
    write_to_excel(course_data, output_excel_filename)
    print('Done')

    print(f'\nGenerating {output_csv_filename}')
    write_to_csv(course_data, output_csv_filename)
    print('Done')


if __name__ == "__main__":
    main()
